"""Debug Spark xlsx parsing — run: python scripts/debug_parse.py [path-to.xlsx]"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import load_workbook

from app.parsers.spark import _find_section_rows, parse_spark_xlsx


def main() -> None:
    if len(sys.argv) < 2:
        root = Path(__file__).resolve().parents[2].parent
        candidates = list(root.glob("*.xlsx"))
        if not candidates:
            print("Usage: python scripts/debug_parse.py file.xlsx")
            sys.exit(1)
        path = candidates[0]
    else:
        path = Path(sys.argv[1])

    print(f"File: {path}")
    data = path.read_bytes()
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    layout_start, absorbance_start = _find_section_rows(ws)
    print(f"Sheet: {ws.title}, rows={ws.max_row}, cols={ws.max_column}")
    print(f"layout_start={layout_start}, absorbance_start={absorbance_start}")

    for row in range(1, min(ws.max_row + 1, 55)):
        a = ws.cell(row, 1).value
        b = ws.cell(row, 2).value
        c = ws.cell(row, 3).value
        if a in ("B", "C", "D", "E", "F", "G") or (isinstance(b, str) and str(b).startswith(("RF", "SM", "BL"))):
            print(f"  row {row}: A={a!r} B={b!r} C={c!r}")

    try:
        result = parse_spark_xlsx(data)
        print(f"Parse OK: {len(result['blocks'])} blocks, {len(result['replicates'])} replicates")
    except Exception as exc:
        print(f"Parse FAILED: {exc}")


if __name__ == "__main__":
    main()
