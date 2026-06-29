"""Parse Tecan Spark MTT plate reader xlsx exports."""

import io
import re
from typing import Any

from openpyxl import load_workbook

PLATE_ROWS = ["B", "C", "D", "E", "F", "G"]
PLATE_ROW_COL = 1  # column A often holds row letter
DATA_COLS = list(range(2, 14))  # B=2 .. M=13


def _cell_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _is_numeric(value: Any) -> bool:
    try:
        float(value)
        return True
    except (TypeError, ValueError):
        return False


def _row_has_well_labels(ws, row_idx: int) -> bool:
    vals = [_cell_str(ws.cell(row_idx, c).value) for c in DATA_COLS]
    return any(v and re.match(r"^(RF|SM|BL)", v, re.IGNORECASE) for v in vals)


def _row_is_absorbance_grid(ws, row_idx: int) -> bool:
    """True if row looks like an absorbance row (>=8 numeric wells in B-M)."""
    numeric = 0
    for c in DATA_COLS:
        if _is_numeric(ws.cell(row_idx, c).value):
            numeric += 1
    return numeric >= 8


def _find_section_rows(ws) -> tuple[int | None, int | None]:
    """Locate well-map header row and absorbance grid start row."""
    layout_start = None
    absorbance_start = None

    for row_idx in range(1, ws.max_row + 1):
        plate_row = _cell_str(ws.cell(row_idx, PLATE_ROW_COL).value)

        if plate_row in PLATE_ROWS and _row_has_well_labels(ws, row_idx):
            if layout_start is None:
                layout_start = row_idx

        if _row_has_well_labels(ws, row_idx) and layout_start is None:
            # Some exports omit row letters in column A
            if row_idx + 5 <= ws.max_row and all(_row_has_well_labels(ws, row_idx + i) for i in range(6)):
                layout_start = row_idx

        if plate_row == "B" and _row_is_absorbance_grid(ws, row_idx):
            if absorbance_start is None:
                absorbance_start = row_idx

    if absorbance_start is None:
        # Fallback: first block of 6 consecutive numeric rows
        for row_idx in range(1, ws.max_row - 4):
            if all(_row_is_absorbance_grid(ws, row_idx + i) for i in range(6)):
                absorbance_start = row_idx
                break

    return layout_start, absorbance_start


def _parse_well_label(label: str) -> dict[str, Any]:
    """Classify well label into control, blank, or dose."""
    label = label.strip()
    upper = label.upper()
    if upper.startswith("RF"):
        return {"role": "control", "dose_index": None, "sample_id": label}
    if upper.startswith("BL"):
        return {"role": "blank", "dose_index": None, "sample_id": label}
    sm_match = re.match(r"SM(\d+)[_-](\d+)", label, re.IGNORECASE)
    if sm_match:
        sm_num = int(sm_match.group(1))
        if 1 <= sm_num <= 10:
            dose_index = sm_num - 1
            block_hint = 1
        elif 11 <= sm_num <= 20:
            dose_index = sm_num - 11
            block_hint = 2
        else:
            dose_index = sm_num - 1
            block_hint = 0
        return {
            "role": "dose",
            "dose_index": dose_index,
            "sample_id": label,
            "sm_number": sm_num,
            "block_hint": block_hint,
        }
    return {"role": "unknown", "dose_index": None, "sample_id": label}


def _read_plate_layout(ws, start_row: int) -> dict[str, dict[int, str]]:
    """Return {plate_row: {col_index: well_label}}."""
    layout: dict[str, dict[int, str]] = {}
    for offset, default_row in enumerate(PLATE_ROWS):
        row_idx = start_row + offset
        if row_idx > ws.max_row:
            break
        plate_row = _cell_str(ws.cell(row_idx, PLATE_ROW_COL).value) or default_row
        if plate_row not in PLATE_ROWS:
            plate_row = default_row
        layout[plate_row] = {}
        for col_idx in DATA_COLS:
            val = _cell_str(ws.cell(row_idx, col_idx).value)
            if val:
                layout[plate_row][col_idx] = val
    return layout


def _read_absorbance_grid(ws, start_row: int) -> dict[str, dict[int, float]]:
    """Return {plate_row: {col_index: absorbance}}."""
    grid: dict[str, dict[int, float]] = {}
    for offset, default_row in enumerate(PLATE_ROWS):
        row_idx = start_row + offset
        if row_idx > ws.max_row:
            break
        plate_row = _cell_str(ws.cell(row_idx, PLATE_ROW_COL).value) or default_row
        if plate_row not in PLATE_ROWS:
            plate_row = default_row
        grid[plate_row] = {}
        for col_idx in DATA_COLS:
            val = ws.cell(row_idx, col_idx).value
            if _is_numeric(val):
                grid[plate_row][col_idx] = float(val)
    return grid


def _build_replicate(layout_row: dict[int, str], abs_row: dict[int, float], plate_row: str) -> dict[str, Any]:
    wells = []
    control_abs = None
    blank_abs = None

    for col_idx, label in layout_row.items():
        if col_idx not in abs_row:
            continue
        meta = _parse_well_label(label)
        well = {
            "well_label": label,
            "column": col_idx,
            "absorbance": abs_row[col_idx],
            "role": meta["role"],
            "dose_index": meta.get("dose_index"),
        }
        wells.append(well)
        if meta["role"] == "control":
            control_abs = abs_row[col_idx]
        elif meta["role"] == "blank":
            blank_abs = abs_row[col_idx]

    return {
        "row_id": f"row_{plate_row}",
        "plate_row": plate_row,
        "wells": wells,
        "control_absorbance": control_abs,
        "blank_absorbance": blank_abs,
    }


def _detect_blocks(replicates: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Group plate rows into compound blocks (B-D block 1, E-G block 2)."""
    blocks_spec = [
        {"block_id": "block_1", "plate_rows": ["B", "C", "D"], "compound_name": "Compound 1", "sm_range": (1, 10)},
        {"block_id": "block_2", "plate_rows": ["E", "F", "G"], "compound_name": "Compound 2", "sm_range": (11, 20)},
    ]

    rep_by_row = {r["plate_row"]: r for r in replicates}
    blocks = []

    for spec in blocks_spec:
        block_reps = [rep_by_row[pr] for pr in spec["plate_rows"] if pr in rep_by_row]
        if not block_reps:
            continue

        dose_count = 0
        for rep in block_reps:
            dose_count = max(dose_count, sum(1 for w in rep["wells"] if w["role"] == "dose"))

        blocks.append(
            {
                "block_id": spec["block_id"],
                "compound_name": spec["compound_name"],
                "plate_rows": spec["plate_rows"],
                "sm_range": spec["sm_range"],
                "n_doses": dose_count,
                "replicates": block_reps,
            }
        )

    # Single-block fallback: treat all rows as one compound
    if not blocks and replicates:
        dose_count = max(sum(1 for w in r["wells"] if w["role"] == "dose") for r in replicates)
        blocks.append(
            {
                "block_id": "block_1",
                "compound_name": "Compound 1",
                "plate_rows": [r["plate_row"] for r in replicates],
                "sm_range": (1, 10),
                "n_doses": dose_count,
                "replicates": replicates,
            }
        )

    return blocks


def _parse_worksheet(ws) -> dict[str, Any]:
    layout_start, absorbance_start = _find_section_rows(ws)
    if layout_start is None or absorbance_start is None:
        raise ValueError(
            "Could not locate plate layout or absorbance grid. "
            "Ensure this is a Tecan Spark export (.xlsx) with RF/SM/BL well labels and absorbance values."
        )

    layout = _read_plate_layout(ws, layout_start)
    absorbance = _read_absorbance_grid(ws, absorbance_start)

    replicates = []
    for plate_row in PLATE_ROWS:
        if plate_row in layout and plate_row in absorbance:
            rep = _build_replicate(layout[plate_row], absorbance[plate_row], plate_row)
            if rep["control_absorbance"] is not None and rep["blank_absorbance"] is not None:
                replicates.append(rep)

    if not replicates:
        raise ValueError(
            "No valid replicate rows with control (RF) and blank (BL) wells found. "
            "Check that the file is a raw Spark export, not a processed workbook."
        )

    blocks = _detect_blocks(replicates)
    if not blocks:
        raise ValueError("Could not detect any compound blocks with dose wells.")

    return {
        "sheet_name": ws.title,
        "layout_start_row": layout_start,
        "absorbance_start_row": absorbance_start,
        "replicates": replicates,
        "blocks": blocks,
    }


def parse_spark_xlsx(file_bytes: bytes) -> dict[str, Any]:
    """Parse a Spark export xlsx and return structured plate data."""
    if file_bytes[:2] != b"PK":
        raise ValueError(
            "This does not look like a valid .xlsx file. "
            "Please export from Spark as Excel (.xlsx), not the older .xls format."
        )

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    errors: list[str] = []

    for sheet in wb.worksheets:
        try:
            return _parse_worksheet(sheet)
        except ValueError as exc:
            errors.append(f"{sheet.title}: {exc}")

    detail = errors[0] if errors else "Unknown format."
    raise ValueError(detail)
