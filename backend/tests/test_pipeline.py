"""Tests for IC50 platform calculations and parser."""

import io
from pathlib import Path

import pytest
from openpyxl import Workbook

from app.calc.dilution import build_concentrations, mg_l_to_um
from app.calc.ic50 import fit_ic50
from app.calc.viability import aggregate_replicates, blank_correct, compute_viability
from app.parsers.spark import parse_spark_xlsx

FIXTURES = Path(__file__).resolve().parents[2].parent.parent


def _make_spark_workbook() -> bytes:
    """Build a minimal Spark-like xlsx in memory."""
    wb = Workbook()
    ws = wb.active

    # Well layout starting row 20
    layout = {
        "B": ["RF1_1"] + [f"SM{i}_1" for i in range(1, 11)] + ["BL1_1"],
        "C": ["RF1_2"] + [f"SM{i}_2" for i in range(1, 11)] + ["BL1_2"],
        "D": ["RF1_3"] + [f"SM{i}_3" for i in range(1, 11)] + ["BL1_3"],
        "E": ["RF1_4"] + [f"SM{i}_1" for i in range(11, 21)] + ["BL1_4"],
        "F": ["RF1_5"] + [f"SM{i}_2" for i in range(11, 21)] + ["BL1_5"],
        "G": ["RF1_6"] + [f"SM{i}_3" for i in range(11, 21)] + ["BL1_6"],
    }
    for i, row in enumerate(["B", "C", "D", "E", "F", "G"]):
        ws.cell(20 + i, 1, row)
        for j, val in enumerate(layout[row]):
            ws.cell(20 + i, 2 + j, val)

    # Absorbance grid starting row 46 — sigmoidal-like response
    doses = [1.5, 1.4, 1.2, 0.9, 0.6, 0.4, 0.25, 0.15, 0.08, 0.05]
    for i, row in enumerate(["B", "C", "D", "E", "F", "G"]):
        ws.cell(46 + i, 1, row)
        ws.cell(46 + i, 2, 1.4)  # control
        for j, d in enumerate(doses):
            ws.cell(46 + i, 3 + j, d + 0.02 * i)
        ws.cell(46 + i, 13, 0.05)  # blank

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_mg_l_to_um():
    # HT karin sheet 2: 50 mg/L, Mw 523.66 -> 95.48 µM
    assert round(mg_l_to_um(50, 523.66), 2) == 95.48


def test_viability_formula():
    # replicate: control=1.4481, blank=0.0562, sample=0.2234
    corr_ctrl = blank_correct(1.4481, 0.0562)
    corr_sample = blank_correct(0.2234, 0.0562)
    viability = compute_viability(corr_sample, corr_ctrl)
    assert viability is not None
    assert round(viability, 1) == 12.0


def test_aggregate_replicates():
    agg = aggregate_replicates([12.0, 14.1, 11.7])
    assert round(agg["mean"], 1) == 12.6
    assert agg["n"] == 3
    assert agg["sd"] is not None


def test_spark_parser_blocks():
    data = _make_spark_workbook()
    result = parse_spark_xlsx(data)
    assert len(result["blocks"]) == 2
    assert result["blocks"][0]["n_doses"] == 10
    assert len(result["blocks"][0]["replicates"]) == 3


def test_ic50_fit_synthetic():
    dose_points = [
        {"um": 100, "mean_viability": 95, "replicates": [{"viability": 95}]},
        {"um": 10, "mean_viability": 90, "replicates": [{"viability": 90}]},
        {"um": 1, "mean_viability": 60, "replicates": [{"viability": 60}]},
        {"um": 0.1, "mean_viability": 20, "replicates": [{"viability": 20}]},
        {"um": 0.01, "mean_viability": 5, "replicates": [{"viability": 5}]},
    ]
    result = fit_ic50(dose_points, n_bootstrap=50)
    assert result["success"]
    assert result["ic50"] is not None
    assert 0.01 < result["ic50"] < 50


@pytest.mark.skipif(
    not Path(__file__).resolve().parents[3].joinpath("Hexa lig-4 rows, dime-2rows.xlsx").exists(),
    reason="Sample Hexa xlsx not in workspace root",
)
def test_parse_hexa_file():
    hexa = Path(__file__).resolve().parents[3] / "Hexa lig-4 rows, dime-2rows.xlsx"
    result = parse_spark_xlsx(hexa.read_bytes())
    assert len(result["blocks"]) >= 1
    assert len(result["replicates"]) >= 3
